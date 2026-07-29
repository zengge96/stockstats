#!/usr/bin/env python3
"""stockstats — A股区间统计工具 (PE/PB百分位分析)"""
import sys, os, argparse, json, urllib.request, statistics, tempfile
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
import akshare as ak

__version__ = "0.1.0"

# ===================== 代码前缀检测 =====================
def detect_prefix(code: str) -> str:
    """
    根据股票代码判断新浪/腾讯前缀。
    6开头 → 'sh' (上证主板)
    0/3开头 → 'sz' (深证主板/创业板)
    5开头 → 'sh' (上证ETF)
    1开头 → 'sz' (深证ETF)
    """
    if code.startswith(('6', '5')):
        return 'sh'
    return 'sz'


# ===================== 财务数据提取 =====================

def is_csi_index(code: str) -> bool:
    """判断是否为中证指数代码(以.CSI结尾)"""
    return code.upper().endswith('.CSI')


def is_etf_fund(code: str) -> bool:
    """判断是否为ETF基金代码(以.ETF结尾)"""
    return code.upper().endswith('.ETF')


# 指数名称 → CSI代码映射(常见指数的硬映射)
INDEX_NAME_MAP = {
    '中证主要消费指数': '000932',
    '中证消费指数': '000932',
    '中证红利指数': '000922',
    '中证红利低波动指数': '930955',
    '中证红利低波100指数': '930955',
    '中证500指数': '000905',
    '中证1000指数': '000852',
    '中证2000指数': '932000',
    '中证800指数': '000906',
    '沪深300指数': '000300',
    '上证50指数': '000016',
    '上证180指数': '000010',
    '科创板50指数': '000688',
    '科创50指数': '000688',
    '深证红利指数': '399324',
    '国证红利指数': '399321',
}


def lookup_etf_index(code: str) -> str:
    """
    通过天天基金网F10页面查ETF跟踪的指数代码。
    返回CSI代码字符串，失败返回None。
    """
    fund_code = code.upper().rstrip('.ETF') if is_etf_fund(code) else code
    import subprocess, re
    try:
        result = subprocess.run([
            'curl', '-s', '--connect-timeout', '10',
            f'https://fundf10.eastmoney.com/jbgk_{fund_code}.html'
        ], capture_output=True, text=True, timeout=15)
        html = result.stdout
        # 方法1: 找跟踪标的表格行
        m = re.search(r'跟踪标的</td><td[^>]*>([^<]+)</td>', html)
        idx_name = m.group(1).strip() if m else None
        if not idx_name:
            # 方法2: 直接搜中证XXX指数
            indices = re.findall(r'中证[^<]{2,20}指数', html)
            # 排除"交易型开放式指数"这种全称中的字段
            idx_name = next((i for i in set(indices) if '交易' not in i), None)
        if idx_name and idx_name in INDEX_NAME_MAP:
            return INDEX_NAME_MAP[idx_name]
        # 方法3: 搜INDEX_NAME_MAP的部分匹配
        if idx_name:
            for name, csi_code in INDEX_NAME_MAP.items():
                if name in idx_name or idx_name in name:
                    return csi_code
    except Exception:
        pass
    return None


def analyze_etf(code: str, years: int = 8) -> dict:
    """分析ETF基金(先查跟踪指数,再用CSIndex取PE)"""
    idx_code = lookup_etf_index(code)
    if not idx_code:
        return {'code': code, 'name': code, 'error': '查不到跟踪指数'}
    # 用CSIndex查底层指数的PE
    try:
        today_str = datetime.now().strftime('%Y%m%d')
        df = ak.stock_zh_index_hist_csindex(symbol=idx_code, end_date=today_str)
    except Exception:
        return {'code': code, 'name': code, 'error': 'CSIndex暂无此指数数据'}
    if df.empty or '滚动市盈率' not in df.columns:
        return {'code': code, 'name': code, 'error': '无历史PE数据'}
    name = df['指数中文简称'].iloc[0]
    pes = df['滚动市盈率'].dropna()
    if len(pes) < 20:
        return {'code': code, 'name': name, 'error': '历史PE数据不足'}
    n = min(years * 250, len(pes))
    recent = pes.tail(n)
    current_pe = float(recent.iloc[-1])
    pe_sorted = sorted(recent)
    pe_percentile = round(sum(1 for p in recent if p < current_pe) / len(recent) * 100, 1)
    date_end = df['日期'].iloc[-1]
    date_start = df['日期'].iloc[-min(n, len(df))] if n < len(df) else df['日期'].iloc[0]
    return {
        'code': code, 'name': name, '统计区间': f'{date_start} ~ {date_end}',
        'trading_days': len(recent),
        'current_pe': round(current_pe, 2), 'pe_percentile': pe_percentile,
        'pe_mean': round(float(recent.mean()), 2), 'pe_median': round(float(recent.median()), 2),
        'pe_high': round(float(recent.max()), 2), 'pe_low': round(float(recent.min()), 2),
        'pe_median_price': None, 'pe_75_price': None, 'pe_25_price': None,
        'current_pb': None, 'pb_percentile': None, 'pb_mean': None, 'pb_median': None,
        'pb_high': None, 'pb_low': None, 'pb_median_price': None, 'pb_75_price': None,
        'pb_25_price': None,
        'dividend_rate': None, 'roe': None,
    }


def analyze_index(code: str, years: int = 8) -> dict:
    """
    用CSIndex分析指数/ETF的PE和百分位。
    返回同 analyze_stock 兼容的字典。
    """
    # 剥离.CSI后缀，传给CSIndex
    idx_code = code.upper().rstrip('.CSI') if is_csi_index(code) else code
    # 从CSIndex获取历史PE数据
    try:
        today_str = datetime.now().strftime("%Y%m%d")
        df = ak.stock_zh_index_hist_csindex(symbol=idx_code, end_date=today_str)
    except Exception:
        return {'code': code, 'name': code, 'error': 'CSIndex暂无此指数数据'}
    
    if df.empty or '滚动市盈率' not in df.columns:
        return {'code': code, 'name': code, 'error': '无历史PE数据'}
    
    name = df['指数中文简称'].iloc[0]
    pes = df['滚动市盈率'].dropna()
    
    if len(pes) < 20:
        return {'code': code, 'name': name, 'error': '历史PE数据不足'}
    
    # 尝试通过新浪指数行情获取最新PE(指数字段33? 不同指数可能不同)
    latest_pe = None
    latest_date = None
    try:
        prefix = 'sh' if idx_code.startswith(('6','9','0')) else 'sz'
        url = f'https://hq.sinajs.cn/list={prefix}{idx_code}'
        req = urllib.request.Request(url, headers={'Referer':'https://finance.sina.com.cn'})
        resp = urllib.request.urlopen(req, timeout=5)
        t = resp.read().decode('gbk')
        parts = t.split(',')
        # 新浪指数field[30]=日期
        if len(parts) > 30:
            d = parts[30].strip().replace('"','')
            if d and d.count('-') == 2:
                latest_date = d
    except Exception:
        pass
    
    # 截取指定年数
    n = min(years * 250, len(pes))
    recent_pes = pes.tail(n)
    
    # 计算日期范围(用实时日期覆盖历史尾值)
    date_end = latest_date if latest_date else df['日期'].iloc[-1]
    date_start = df['日期'].iloc[-min(n, len(df))] if n < len(df) else df['日期'].iloc[0]
    
    # 当前PE仍用历史数据最后值(因为没有PE字段)
    current_pe = float(recent_pes.iloc[-1])
    pe_sorted = sorted(recent_pes)
    pe_percentile = round(sum(1 for p in recent_pes if p < current_pe) / len(recent_pes) * 100, 1)
    
    result = {
        'code': code,
        'name': name,
        '统计区间': f'{date_start} ~ {date_end}',
        'trading_days': len(recent_pes),
        'current_pe': round(current_pe, 2),
        'pe_percentile': pe_percentile,
        'pe_mean': round(float(recent_pes.mean()), 2),
        'pe_median': round(float(recent_pes.median()), 2),
        'pe_high': round(float(recent_pes.max()), 2),
        'pe_low': round(float(recent_pes.min()), 2),
        # ETF没有EPS/NAV,无法算分位价和PB
        'pe_median_price': None, 'pe_75_price': None, 'pe_25_price': None,
        'current_pb': None, 'pb_percentile': None,
        'pb_mean': None, 'pb_median': None, 'pb_high': None, 'pb_low': None,
        'pb_median_price': None, 'pb_75_price': None, 'pb_25_price': None,
        'dividend_rate': None, 'roe': None,
    }
    return result


def get_annual_eps(code: str) -> dict:
    """
    从 stock_financial_abstract 提取年报EPS。
    返回 {年份: EPS, ...} 字典。
    """
    try:
        df = ak.stock_financial_abstract(symbol=code)
    except Exception:
        return None
    
    annual_cols = sorted(
        [c for c in df.columns
         if str(c).isdigit() and str(c)[4:8] == '1231']
    )
    eps_row = df[df['指标'] == '基本每股收益']
    if eps_row.empty:
        return None
    
    eps_map = {}
    for c in annual_cols:
        v = eps_row.iloc[0][c]
        if v is not None and not (isinstance(v, float) and (v != v or v == 0)):
            eps_map[int(str(c)[:4])] = float(v)
    return eps_map if eps_map else None


def get_annual_nav(code: str) -> dict:
    """
    提取年报每股净资产。
    返回 {年份: 每股净资产, ...} 字典。
    """
    try:
        df = ak.stock_financial_abstract(symbol=code)
    except Exception:
        return None
    
    annual_cols = sorted(
        [c for c in df.columns
         if str(c).isdigit() and str(c)[4:8] == '1231']
    )
    nav_row = df[df['指标'] == '每股净资产']
    if nav_row.empty:
        return None
    
    nav_map = {}
    for c in annual_cols:
        v = nav_row.iloc[0][c]
        if v is not None and not (isinstance(v, float) and (v != v or v == 0)):
            nav_map[int(str(c)[:4])] = float(v)
    return nav_map if nav_map else None


def get_latest_roe(code: str) -> float:
    """提取最新年报ROE(%)"""
    try:
        df = ak.stock_financial_abstract(symbol=code)
    except Exception:
        return None
    
    annual_cols = sorted(
        [c for c in df.columns
         if str(c).isdigit() and str(c)[4:8] == '1231']
    )
    # akshare的指标名是'净资产收益率(ROE)'或'加权净资产收益率'
    roe_row = df[df['指标'] == '净资产收益率(ROE)']
    if roe_row.empty:
        roe_row = df[df['指标'] == '加权净资产收益率']
    if roe_row.empty:
        return None
    
    v = roe_row.iloc[0][annual_cols[-1]]
    if v is not None and not (isinstance(v, float) and (v != v or v == 0)):
        return float(v)
    return None


def get_current_dividend_rate(code: str) -> float:
    """从腾讯实时行情获取当前TTM股息率 field[64]"""
    prefix = detect_prefix(code)
    try:
        url = f"http://qt.gtimg.cn/q={prefix}{code}"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as r:
            raw = r.read().decode('gbk')
        parts = raw.split('~')
        if len(parts) > 64 and parts[64]:
            return float(parts[64])
    except Exception:
        pass
    return None


# ===================== PE/PB计算 =====================
def get_price_history(code: str, prefix: str, max_days: int = 2500) -> list:
    """获取日线数据（新浪API）"""
    url = (
        f"https://money.finance.sina.com.cn/quotes_service/api/json_v2.php/"
        f"CN_MarketData.getKLineData?symbol={prefix}{code}&scale=240&ma=no&datalen={max_days}"
    )
    try:
        with urllib.request.urlopen(url, timeout=10) as resp:
            return json.loads(resp.read().decode('utf-8'))
    except Exception:
        return []


def get_stock_name(code: str, prefix: str) -> str:
    """获取股票名称"""
    try:
        url = f"http://qt.gtimg.cn/q={prefix}{code}"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as r:
            raw = r.read().decode('gbk')
        parts = raw.split('~')
        if len(parts) > 1:
            return parts[1]
    except Exception:
        pass
    return code


def compute_pe_pb_history(daily_all: list, eps_map: dict, nav_map: dict,
                          years: int = 8, block_years: set = None):
    """
    计算PE和PB历史序列 (年报法: 前一年EPS逐年代入)。
    返回 (pe_history, pb_history, date_range)
    """
    pe_history = []
    pb_history = []
    date_range = None
    
    if block_years is None:
        block_years = set()
    
    for d in daily_all:
        day_str = d['day']
        year = int(day_str[:4])
        close = float(d['close'])
        
        # 屏蔽区间
        if year in block_years:
            continue
        
        # PE: 前一年EPS
        known_eps = eps_map.get(year - 1) or eps_map.get(year)
        if known_eps and known_eps > 0 and 5 < close / known_eps < 80:
            pe_history.append(close / known_eps)
        
        # PB: 前一年NAV
        known_nav = nav_map.get(year - 1) or nav_map.get(year)
        if known_nav and known_nav > 0 and 0.5 < close / known_nav < 20:
            pb_history.append(close / known_nav)
    
    if date_range is None and len(daily_all) >= 2:
        date_range = f"{daily_all[0]['day']} ~ {daily_all[-1]['day']}"
    
    return pe_history, pb_history, date_range


def analyze_stock(code: str, years: int = 8, block_years: set = None) -> dict:
    """
    分析单只股票，返回全部统计指标。
    
    返回:
    {
        'code': '600036',
        'name': '招商银行',
        '统计区间': '...',
        'trading_days': 2000,
        'current_pe': 6.8,
        'pe_percentile': 20.4,
        'pe_mean': 9.0,
        'pe_median': 8.5,
        'pe_high': 15.0,
        'pe_low': 5.0,
        'pe_median_price': ...,
        'pe_75_price': ...,
        'pe_25_price': ...,
        'current_pb': 0.89,
        ... (PB同理)
        'dividend_rate': 5.16,
        'roe': 13.71,
    }
    """
    if block_years is None:
        block_years = set()
    
    prefix = detect_prefix(code)
    name = get_stock_name(code, prefix)
    
    # 财务数据
    eps_map = get_annual_eps(code)
    nav_map = get_annual_nav(code)
    roe = get_latest_roe(code)
    dividend_rate = get_current_dividend_rate(code)
    
    if eps_map is None or nav_map is None:
        return {'code': code, 'name': name, 'error': '财务数据获取失败'}
    
    latest_eps = eps_map.get(max(eps_map.keys()))
    latest_nav = nav_map.get(max(nav_map.keys()))
    
    # 日线
    daily_all = get_price_history(code, prefix)
    if not daily_all:
        return {'code': code, 'name': name, 'error': '日线数据获取失败'}
    
    current_close = float(daily_all[-1]['close'])
    
    # 截取指定年数
    # 年 ≈ 250 个交易日
    n_days = min(years * 250, len(daily_all))
    daily_subset = daily_all[-n_days:]
    
    # 计算PE/PB历史
    pe_history, pb_history, date_range = compute_pe_pb_history(
        daily_subset, eps_map, nav_map, years, block_years
    )
    
    trading_days = len(pe_history)
    
    # === PE统计 ===
    current_pe = current_close / latest_eps if latest_eps and latest_eps > 0 else None
    
    pe_percentile = None
    pe_mean = None
    pe_median = None
    pe_high = None
    pe_low = None
    pe_median_price = None
    pe_75_price = None
    pe_25_price = None
    
    if pe_history and current_pe:
        pe_sorted = sorted(pe_history)
        pe_percentile = round(sum(1 for p in pe_history if p < current_pe) / len(pe_history) * 100, 1)
        pe_mean = round(statistics.mean(pe_history), 1)
        pe_median = round(statistics.median(pe_history), 1)
        pe_high = round(max(pe_history), 1)
        pe_low = round(min(pe_history), 1)
        
        # PE分位价
        def pe_quantile_price(pct):
            idx = int(len(pe_sorted) * pct / 100)
            idx = min(idx, len(pe_sorted) - 1)
            return round(pe_sorted[idx] * latest_eps, 2) if latest_eps else None
        
        pe_median_price = round(pe_median * latest_eps, 2) if latest_eps else None
        pe_75_price = pe_quantile_price(75)
        pe_25_price = pe_quantile_price(25)
    
    # === PB统计 ===
    current_pb = current_close / latest_nav if latest_nav and latest_nav > 0 else None
    
    pb_percentile = None
    pb_mean = None
    pb_median = None
    pb_high = None
    pb_low = None
    pb_median_price = None
    pb_75_price = None
    pb_25_price = None
    
    if pb_history and current_pb:
        pb_sorted = sorted(pb_history)
        pb_percentile = round(sum(1 for p in pb_history if p < current_pb) / len(pb_history) * 100, 1)
        pb_mean = round(statistics.mean(pb_history), 2)
        pb_median = round(statistics.median(pb_history), 2)
        pb_high = round(max(pb_history), 2)
        pb_low = round(min(pb_history), 2)
        
        def pb_quantile_price(pct):
            idx = int(len(pb_sorted) * pct / 100)
            idx = min(idx, len(pb_sorted) - 1)
            return round(pb_sorted[idx] * latest_nav, 2) if latest_nav else None
        
        pb_median_price = round(pb_median * latest_nav, 2) if latest_nav else None
        pb_75_price = pb_quantile_price(75)
        pb_25_price = pb_quantile_price(25)
    
    return {
        'code': code,
        'name': name,
        '统计区间': date_range or '',
        'trading_days': trading_days,
        'current_pe': round(current_pe, 2) if current_pe else None,
        'pe_percentile': pe_percentile,
        'pe_mean': pe_mean,
        'pe_median': pe_median,
        'pe_high': pe_high,
        'pe_low': pe_low,
        'pe_median_price': pe_median_price,
        'pe_75_price': pe_75_price,
        'pe_25_price': pe_25_price,
        'current_pb': round(current_pb, 2) if current_pb else None,
        'pb_percentile': pb_percentile,
        'pb_mean': pb_mean,
        'pb_median': pb_median,
        'pb_high': pb_high,
        'pb_low': pb_low,
        'pb_median_price': pb_median_price,
        'pb_75_price': pb_75_price,
        'pb_25_price': pb_25_price,
        'dividend_rate': dividend_rate,
        'roe': round(roe, 2) if roe else None,
    }


# ===================== 多股票 =====================
def analyze_multiple(codes: list, years: int = 8, block_years: set = None) -> list:
    results = []
    for c in codes:
        if is_etf_fund(c):
            results.append(analyze_etf(c, years))
        elif is_csi_index(c):
            results.append(analyze_index(c.upper(), years))
        else:
            results.append(analyze_stock(c, years, block_years))
    return results


# ===================== Excel报告 =====================
def generate_report(codes: list, years: int = 8,
                     block_years: set = None, output: str = None):
    """生成xlsx报告"""
    results = analyze_multiple(codes, years, block_years)
    
    if output is None:
        code_str = '_'.join(c[:6] for c in codes[:5])
        output = f'{code_str}.xlsx'
    
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    # Sheet 1: 统计总表
    columns = [
        '代码', '名称', '统计区间', '交易日数',
        'TTM PE', 'PE百分位', 'PE均值', 'PE中位', 'PE最高', 'PE最低',
        'PE中位数价', 'PE75%分位价', 'PE25%分位价',
        'TTM PB', 'PB百分位', 'PB均值', 'PB中位', 'PB最高', 'PB最低',
        'PB中位数价', 'PB75%分位价', 'PB25%分位价',
        'TTM股息率[TTM]', 'ROE'
    ]
    
    rows = []
    for r in results:
        if 'error' in r:
            rows.append([r.get('code'), r.get('name'), 'ERROR: ' + r['error']] +
                        [''] * (len(columns) - 3))
        else:
            rows.append([
                r['code'], r['name'], r['统计区间'], r['trading_days'],
                r['current_pe'], r['pe_percentile'], r['pe_mean'], r['pe_median'],
                r['pe_high'], r['pe_low'], r['pe_median_price'], r['pe_75_price'],
                r['pe_25_price'],
                r['current_pb'], r['pb_percentile'], r['pb_mean'], r['pb_median'],
                r['pb_high'], r['pb_low'], r['pb_median_price'], r['pb_75_price'],
                r['pb_25_price'],
                r['dividend_rate'], r['roe'],
            ])
    
    df_summary = pd.DataFrame(rows, columns=columns)
    # 代码列保存为字符串，避免pandas转为int
    df_summary['代码'] = df_summary['代码'].astype(str)
    df_summary.to_excel(writer, sheet_name='统计总表', index=False)
    
    # Sheet 2+: 每个代码原始数据(用Excel公式计算PE/PB)
    for code in codes:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        sheet_name = f'原始_{code}'
        ws = writer.book.create_sheet(title=sheet_name[:31])
        
        prefix = detect_prefix(code)
        daily_all = get_price_history(code, prefix)
        is_idx_or_etf = is_csi_index(code) or is_etf_fund(code)
        eps_map = get_annual_eps(code) if not is_idx_or_etf else None
        nav_map = get_annual_nav(code) if not is_idx_or_etf else None
        
        if is_idx_or_etf:
            # === 指数/ETF分支: 用CSIndex原始PE数据 ===
            try:
                today_str = datetime.now().strftime('%Y%m%d')
                if is_etf_fund(code):
                    idx_code = lookup_etf_index(code) or code
                else:
                    idx_code = code.rstrip('.CSI').rstrip('.csi')
                df_idx = ak.stock_zh_index_hist_csindex(symbol=idx_code, end_date=today_str)
            except Exception:
                continue
            
            if df_idx.empty or '滚动市盈率' not in df_idx.columns:
                continue
            
            headers = ['日期', '指数收盘', '滚动市盈率']
            for j, h in enumerate(headers):
                cell = ws.cell(row=1, column=j+1)
                cell.value = h
                cell.font = openpyxl.styles.Font(bold=True)
            
            row_num = 2
            for _, row in df_idx.iterrows():
                date_obj = row['日期']
                if hasattr(date_obj, 'strftime'):
                    cell_date = ws.cell(row=row_num, column=1, value=date_obj)
                    cell_date.number_format = 'yyyy-mm-dd'
                ws.cell(row=row_num, column=2, value=row.get('收盘', 0))
                ws.cell(row=row_num, column=3, value=row.get('滚动市盈率', ''))
                row_num += 1
            
            # PE趋势图
            d3 = Reference(ws, min_col=3, min_row=1, max_row=row_num-1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=row_num-1)
            c = LineChart()
            c.title = 'PE走势'
            c.add_data(d3, titles_from_data=True)
            c.set_categories(cats)
            c.style = 2
            ws.add_chart(c, "E1")
            
        elif eps_map and daily_all:
            # === 个股分支: 价格+EPS公式算PE ===
            headers = ['日期', '收盘价', 'EPS', '每股净资产', '成交量', 'PE', 'PB']
            for j, h in enumerate(headers):
                cell = ws.cell(row=1, column=j+1)
                cell.value = h
                cell.font = openpyxl.styles.Font(bold=True)
            
            row_num = 2
            for d in daily_all[-2500:]:
                close = float(d['close'])
                yr = int(d['day'][:4])
                known_eps = eps_map.get(yr - 1) or eps_map.get(yr)
                known_nav = nav_map.get(yr - 1) or nav_map.get(yr)
                if not known_eps or known_eps <= 0:
                    continue
                date_obj = datetime.strptime(d['day'], '%Y-%m-%d').date()
                cell_date = ws.cell(row=row_num, column=1, value=date_obj)
                cell_date.number_format = 'yyyy-mm-dd'
                ws.cell(row=row_num, column=2, value=close)
                ws.cell(row=row_num, column=3, value=known_eps)
                ws.cell(row=row_num, column=4, value=known_nav if known_nav else '')
                ws.cell(row=row_num, column=5, value=float(d.get('volume', 0)))
                ws.cell(row=row_num, column=6).value = f'=IF(C{row_num}>0,B{row_num}/C{row_num},NA())'
                ws.cell(row=row_num, column=7).value = f'=IF(D{row_num}>0,B{row_num}/D{row_num},NA())'
                row_num += 1
            
            # 3个图表
            d2 = Reference(ws, min_col=2, min_row=1, max_row=row_num-1)
            d6 = Reference(ws, min_col=6, min_row=1, max_row=row_num-1)
            d7 = Reference(ws, min_col=7, min_row=1, max_row=row_num-1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=row_num-1)
            c1 = LineChart(); c1.title = '股价走势'; c1.add_data(d2, titles_from_data=True); c1.set_categories(cats)
            c2 = LineChart(); c2.title = 'PE走势'; c2.add_data(d6, titles_from_data=True); c2.set_categories(cats)
            c3 = LineChart(); c3.title = 'PB走势'; c3.add_data(d7, titles_from_data=True); c3.set_categories(cats)
            c1.style = 2; c2.style = 2; c3.style = 2
            ws.add_chart(c1, "I1"); ws.add_chart(c2, "I19"); ws.add_chart(c3, "I37")
        else:
            # 无数据
            ws.cell(row=1, column=1, value='无原始数据')
        
    writer.close()
    return output


# ===================== CLI入口 =====================
def parse_args(argv=None):
    parser = argparse.ArgumentParser(
        description='stockstats — A股区间统计工具 (PE/PB百分位分析)'
    )
    parser.add_argument('codes', nargs='+', help='股票代码，空格分隔')
    parser.add_argument('-y', '--years', type=int, default=8,
                        help='统计年数，默认8，范围1-10')
    parser.add_argument('-b', '--block', action='append', default=[],
                        help='屏蔽年份(可重复)，如 -b 2021 -b 2022')
    parser.add_argument('-o', '--output', type=str, default=None,
                        help='输出xlsx文件名')
    return parser.parse_args(argv)


def parse_block_years(block_args: list) -> set:
    """解析 -b 参数为年份集合"""
    years = set()
    for arg in block_args:
        if '-' in arg:
            parts = arg.split('-')
            start, end = int(parts[0]), int(parts[1])
            years.update(range(start, end + 1))
        else:
            years.add(int(arg))
    return years


def main():
    args = parse_args()
    block_years = parse_block_years(args.block)
    
    print(f"stockstats v{__version__}")
    print(f"代码: {' '.join(args.codes)}")
    print(f"周期: {args.years}年")
    if block_years:
        print(f"屏蔽: {sorted(block_years)}")
    print()
    
    results = analyze_multiple(args.codes, args.years, block_years)
    
    # 控制台输出统计总表
    print(f"{'代码':>6} {'名称':<10} {'PE':>6} {'P%':>6} {'均值':>6} {'PB':>6} {'ROE':>5} {'股息率':>6}")
    print("-" * 55)
    for r in results:
        if 'error' in r:
            print(f"{r.get('code','?'):>6} {r.get('name','?'):<10} ❌ {r.get('error','?')}")
        else:
            pe = f"{r.get('current_pe',0):.1f}x" if r.get('current_pe') else 'N/A'
            pp = f"P{r.get('pe_percentile',0):.0f}" if r.get('pe_percentile') is not None else 'N/A'
            pm = f"{r.get('pe_mean',0):.1f}x" if r.get('pe_mean') else 'N/A'
            pb = f"{r.get('current_pb',0):.1f}x" if r.get('current_pb') else 'N/A'
            ro = f"{r.get('roe',0):.1f}%" if r.get('roe') else 'N/A'
            dr = f"{r.get('dividend_rate',0):.1f}%" if r.get('dividend_rate') else 'N/A'
            print(f"{r['code']:>6} {r['name']:<10} {pe:>6} {pp:>6} {pm:>6} {pb:>6} {ro:>5} {dr:>6}")
    print()
    
    out_path = generate_report(args.codes, years=args.years,
        block_years=block_years, output=args.output
    )
    print(f"✅ 报告已生成: {os.path.abspath(out_path)}")


if __name__ == '__main__':
    main()